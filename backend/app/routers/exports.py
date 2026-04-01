"""Feature 1.6 — Excel Export Endpoints."""
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import Auction, Item, User
from app.fees import calculate_fees
from app.calculators import calculate_ebay_fees
from app.middleware.limits import check_export_permission, raise_http_error_from_limit_error, LimitError
from app.auth.jwt import require_auth
router = APIRouter(prefix="/api/export", tags=["export"])


def _get_item_fees(item):
    """Get platform fees for an item based on its channel."""
    channel = item.sale_channel or "ebay"
    if not item.sold_price:
        return 0.0
    if channel == "ebay":
        ebay = calculate_ebay_fees(item.sold_price, item.shipping_charged_buyer, promoted_pct=item.promoted_pct)
        return ebay["total_ebay_fees"]
    result = calculate_fees(item.sold_price, channel, item.shipping_cost_actual, item.buy_cost_total)
    return result["platform_fee"]


def _make_xlsx(headers, rows, sheet_name="Sheet1"):
    """Create an in-memory Excel file with formatted headers."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    bold = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)

    # Auto-width
    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 40)

    # Freeze top row
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


ITEM_EXPORT_COLUMNS = [
    "Item Name", "Category", "Lot #", "Hammer Price", "Total Buy Cost",
    "Sale Channel", "Platform Sold On", "List Price", "Sold Price", "Est. Resale", "Date Sold",
    "Your Shipping Cost", "Buyer Shipping Charge", "Promoted Listing %",
    "Total Platform Fees", "Net Profit", "ROI %",
    "Status", "Notes",
]


def _item_to_row(item):
    fees = _get_item_fees(item) if item.sold_price else 0.0
    return [
        item.name, item.category, item.lot_number, item.hammer_price, item.buy_cost_total,
        item.sale_channel or "ebay", item.platform_sold_on or item.sale_channel or "ebay",
        item.list_price, item.sold_price, item.estimated_resale, item.sell_date,
        item.shipping_cost_actual, item.shipping_charged_buyer, item.promoted_pct,
        round(fees, 2), item.net_profit, item.roi_pct,
        item.status, item.notes,
    ]


@router.get("/auction/{auction_id}")
def export_auction(auction_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    # First check if auction exists and belongs to user
    auction = db.query(Auction).filter(Auction.id == auction_id, Auction.user_id == current_user.id).first()
    if not auction:
        raise HTTPException(status_code=404, detail="Auction not found")
    
    # Check export permission
    try:
        check_export_permission(current_user.plan)
    except LimitError as e:
        raise_http_error_from_limit_error(e)
    items = db.query(Item).filter(Item.auction_id == auction_id).all()
    if not items:
        rows = [["No items in this auction"] + [""] * (len(ITEM_EXPORT_COLUMNS) - 1)]
    else:
        rows = [_item_to_row(i) for i in items]
    buf = _make_xlsx(ITEM_EXPORT_COLUMNS, rows, f"Auction {auction.name}")
    safe_name = auction.name.replace(" ", "_").replace("/", "-")
    filename = f"flipiq_auction_{safe_name}_{auction.date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/inventory")
def export_inventory(start: str = None, end: str = None, channel: str = None, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    # Check export permission
    try:
        check_export_permission(current_user.plan)
    except LimitError as e:
        raise_http_error_from_limit_error(e)
    
    query = db.query(Item).join(Auction).filter(Auction.user_id == current_user.id)
    if start:
        query = query.filter(Auction.date >= start)
    if end:
        query = query.filter(Auction.date <= end)
    if channel and channel != "all":
        query = query.filter(Item.sale_channel == channel)
    items = query.all()
    rows = [_item_to_row(i) for i in items]
    buf = _make_xlsx(ITEM_EXPORT_COLUMNS, rows, "Inventory")
    from datetime import date
    filename = f"flipiq_inventory_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/dashboard/summary")
def export_dashboard_summary(year: int = 2026, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    # Check export permission
    try:
        check_export_permission(current_user.plan)
    except LimitError as e:
        raise_http_error_from_limit_error(e)
    
    items = db.query(Item).join(Auction).filter(
        Auction.date.like(f"{year}%"),
        Auction.user_id == current_user.id
    ).all()
    monthly = {}
    for item in items:
        auction = db.query(Auction).filter(Auction.id == item.auction_id).first()
        if not auction:
            continue
        month = auction.date[:7]
        if month not in monthly:
            monthly[month] = {"month": month, "revenue": 0, "cogs": 0, "fees": 0, "profit": 0}
        monthly[month]["cogs"] += item.buy_cost_total
        if item.status == "sold" and item.sold_price:
            monthly[month]["revenue"] += item.sold_price
            monthly[month]["fees"] += _get_item_fees(item)
            monthly[month]["profit"] += (item.net_profit or 0)

    headers = ["Month", "Revenue", "COGS", "Platform Fees", "Net Profit"]
    rows = []
    for m in sorted(monthly.values(), key=lambda x: x["month"]):
        rows.append([m["month"], round(m["revenue"], 2), round(m["cogs"], 2), round(m["fees"], 2), round(m["profit"], 2)])

    buf = _make_xlsx(headers, rows, f"P&L Summary {year}")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="flipiq_dashboard_summary_{year}.xlsx"'},
    )


@router.get("/tax-summary")
def export_tax_summary(year: int = 2026, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    # Check export permission
    try:
        check_export_permission(current_user.plan)
    except LimitError as e:
        raise_http_error_from_limit_error(e)
    
    items = db.query(Item).join(Auction).filter(
        Auction.date.like(f"{year}%"),
        Item.status == "sold",
        Auction.user_id == current_user.id
    ).all()
    channels = {}
    for item in items:
        ch = item.sale_channel or "ebay"
        if ch not in channels:
            channels[ch] = {"channel": ch, "revenue": 0, "cogs": 0, "fees": 0, "profit": 0}
        channels[ch]["revenue"] += (item.sold_price or 0)
        channels[ch]["cogs"] += item.buy_cost_total
        channels[ch]["fees"] += _get_item_fees(item)
        channels[ch]["profit"] += (item.net_profit or 0)

    headers = ["Channel", "Total Revenue", "Total COGS", "Total Fees", "Net Profit"]
    rows = []
    totals = [0, 0, 0, 0]
    for c in sorted(channels.values(), key=lambda x: x["revenue"], reverse=True):
        rows.append([c["channel"], round(c["revenue"], 2), round(c["cogs"], 2), round(c["fees"], 2), round(c["profit"], 2)])
        totals[0] += c["revenue"]
        totals[1] += c["cogs"]
        totals[2] += c["fees"]
        totals[3] += c["profit"]
    rows.append(["TOTAL", round(totals[0], 2), round(totals[1], 2), round(totals[2], 2), round(totals[3], 2)])

    buf = _make_xlsx(headers, rows, f"Tax Summary {year}")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="flipiq_tax_summary_{year}.xlsx"'},
    )
