"""Phase 1 tests — Auction House Configs, Province Tax, Break-Even Alerts, Export, Currency, Goal Tracker."""
import pytest
from io import BytesIO

from fastapi.testclient import TestClient
from app.main import app

client = TestClient(app)

# ── Feature 1.1 — Auction House Configs ─────────────────────────────────────

from app.buy_cost import (
    calculate_buy_cost, get_preset_config, check_price_alert,
    CANADA_TAX_RATES, PRESET_AUCTION_HOUSES, SHIPPING_PRESETS,
)


class TestAuctionHouseConfig:
    def test_encore_buy_cost_etransfer(self):
        config = get_preset_config("Encore Auctions (London ON)")
        result = calculate_buy_cost(50.00, config, "etransfer")
        # hammer=50, premium=8, handling=1.50, subtotal=59.50, tax=7.735→7.74, total=67.235→67.24
        assert result["total_buy_cost"] == pytest.approx(67.24, abs=0.02)

    def test_encore_buy_cost_credit_card(self):
        config = get_preset_config("Encore Auctions (London ON)")
        result = calculate_buy_cost(50.00, config, "credit_card")
        assert result["credit_card_surcharge"] > 0
        assert result["total_buy_cost"] > 67.20

    def test_maxsold_no_premium(self):
        config = get_preset_config("MaxSold")
        result = calculate_buy_cost(50.00, config, "credit_card")
        assert result["buyer_premium"] == 0.00
        # MaxSold taxes hammer_only: tax = 50 * 0.13 = 6.5, total = 56.50
        assert result["tax"] == 6.50

    def test_custom_config_zero_fees(self):
        config = {
            "buyer_premium_pct": 0, "handling_fee_mode": "none",
            "handling_fee_flat": 0, "handling_fee_pct": 0,
            "tax_pct": 0, "tax_applies_to": "subtotal",
            "credit_card_surcharge_pct": 0, "online_bidding_fee_pct": 0,
        }
        result = calculate_buy_cost(50.00, config, "cash")
        assert result["total_buy_cost"] == 50.00

    def test_hibid_credit_card_surcharge(self):
        config = get_preset_config("HiBid (Generic)")
        result = calculate_buy_cost(100.00, config, "credit_card")
        assert result["credit_card_surcharge"] > 0
        assert result["buyer_premium"] == 15.00  # 15% of 100

    def test_bidfta_handling_fee(self):
        config = get_preset_config("BidFTA")
        result = calculate_buy_cost(100.00, config, "etransfer")
        assert result["handling_fee"] == 3.00

    def test_ritchie_bros_surcharge(self):
        config = get_preset_config("Ritchie Bros. / IronPlanet")
        result = calculate_buy_cost(100.00, config, "credit_card")
        assert result["credit_card_surcharge"] > 0
        assert result["buyer_premium"] == 10.00

    def test_all_presets_exist(self):
        assert len(PRESET_AUCTION_HOUSES) == 6

    def test_unknown_preset_raises(self):
        with pytest.raises(ValueError):
            get_preset_config("NonExistent Auction")

    def test_pct_handling_mode(self):
        config = {
            "buyer_premium_pct": 0, "handling_fee_mode": "pct",
            "handling_fee_flat": 0, "handling_fee_pct": 5.0,
            "tax_pct": 0, "tax_applies_to": "subtotal",
            "credit_card_surcharge_pct": 0, "online_bidding_fee_pct": 0,
        }
        result = calculate_buy_cost(100.00, config, "cash")
        assert result["handling_fee"] == 5.00

    def test_online_bidding_fee(self):
        config = {
            "buyer_premium_pct": 0, "handling_fee_mode": "none",
            "handling_fee_flat": 0, "handling_fee_pct": 0,
            "tax_pct": 0, "tax_applies_to": "subtotal",
            "credit_card_surcharge_pct": 0, "online_bidding_fee_pct": 3.0,
        }
        result = calculate_buy_cost(100.00, config, "cash")
        assert result["online_bidding_fee"] == 3.00


# ── Feature 1.2 — Province Tax ──────────────────────────────────────────────

class TestProvinceTax:
    def test_ontario_tax_rate(self):
        assert CANADA_TAX_RATES["ON"]["rate"] == 13.0

    def test_alberta_gst_only(self):
        assert CANADA_TAX_RATES["AB"]["rate"] == 5.0
        assert CANADA_TAX_RATES["AB"]["type"] == "GST only"

    def test_quebec_rate(self):
        assert CANADA_TAX_RATES["QC"]["rate"] == 14.975

    def test_all_13_provinces_present(self):
        assert len(CANADA_TAX_RATES) == 13

    def test_all_have_required_fields(self):
        for code, info in CANADA_TAX_RATES.items():
            assert "name" in info
            assert "rate" in info
            assert "type" in info
            assert isinstance(info["rate"], (int, float))


# ── Feature 1.3 — Shipping Presets ──────────────────────────────────────────

class TestShippingPresets:
    def test_presets_exist(self):
        assert len(SHIPPING_PRESETS) == 9

    def test_local_pickup_zero(self):
        local = [p for p in SHIPPING_PRESETS if "Local Pickup" in p["name"]][0]
        assert local["cost_cad"] == 0.00

    def test_all_have_cost(self):
        for p in SHIPPING_PRESETS:
            assert "cost_cad" in p
            assert isinstance(p["cost_cad"], (int, float))


# ── Feature 1.5 — Break-Even Alerts ─────────────────────────────────────────

class TestBreakEvenAlerts:
    def test_below_buy_cost_triggers_red(self):
        alert = check_price_alert(sell_price=5.00, buy_cost=10.00, platform_fees=2.00, target_margin=30.0)
        assert alert["level"] == "red"

    def test_below_fees_triggers_orange(self):
        alert = check_price_alert(sell_price=11.00, buy_cost=10.00, platform_fees=2.00, target_margin=30.0)
        assert alert["level"] == "orange"

    def test_below_target_margin_triggers_yellow(self):
        alert = check_price_alert(sell_price=14.00, buy_cost=10.00, platform_fees=2.00, target_margin=30.0)
        assert alert["level"] == "yellow"

    def test_healthy_margin_triggers_green(self):
        alert = check_price_alert(sell_price=20.00, buy_cost=10.00, platform_fees=2.00, target_margin=30.0)
        assert alert["level"] == "green"

    def test_red_alert_has_message(self):
        alert = check_price_alert(sell_price=5.00, buy_cost=10.00, platform_fees=2.00)
        assert "LOSE" in alert["message"]

    def test_green_no_target_margin(self):
        alert = check_price_alert(sell_price=20.00, buy_cost=10.00, platform_fees=2.00, target_margin=0.0)
        assert alert["level"] == "green"

    def test_zero_buy_cost_safe(self):
        alert = check_price_alert(sell_price=10.00, buy_cost=0.00, platform_fees=0.00)
        assert alert["level"] == "green"


# ── Feature 1.7 — Currency ──────────────────────────────────────────────────

class TestCurrency:
    def test_cad_to_usd_conversion(self):
        rate = 0.73
        assert round(100.00 * rate, 2) == 73.00

    def test_conversion_roundtrip(self):
        rate = 0.73
        cad = 100.0
        usd = cad * rate
        back = usd / rate
        assert round(back, 2) == 100.0


# ── Feature 1.8 — Profit Goal ───────────────────────────────────────────────

def calculate_goal_progress(current: float, goal: float) -> dict:
    """Standalone test helper matching dashboard logic."""
    if goal <= 0:
        return {"enabled": False}
    remaining = max(0.0, goal - current)
    pct = round(current / goal * 100, 1)
    return {
        "enabled": True,
        "pct": pct,
        "remaining": remaining,
        "exceeded": current >= goal,
        "overage": round(current - goal, 2) if current > goal else 0,
    }


class TestProfitGoal:
    def test_goal_progress_calculation(self):
        progress = calculate_goal_progress(current=1340.00, goal=2000.00)
        assert progress["pct"] == 67.0
        assert progress["remaining"] == 660.00

    def test_goal_exceeded_shows_overage(self):
        progress = calculate_goal_progress(current=2340.00, goal=2000.00)
        assert progress["exceeded"] == True
        assert progress["overage"] == 340.00

    def test_zero_goal_disables_tracker(self):
        progress = calculate_goal_progress(current=500.00, goal=0.0)
        assert progress["enabled"] == False


# ── Feature 1.6 — Export (API tests) ────────────────────────────────────────

class TestExport:
    def test_export_auction_returns_xlsx(self):
        resp = client.post("/api/auctions/", json={"name": "Export Test", "date": "2026-03-25", "total_hammer": 100.0})
        aid = resp.json()["id"]
        client.post(f"/api/auctions/{aid}/items", json={
            "name": "Widget", "hammer_price": 30.0, "status": "sold",
            "sold_price": 50.0, "sell_date": "2026-03-26", "sale_channel": "ebay",
        })
        response = client.get(f"/api/export/auction/{aid}")
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

    def test_export_auction_not_found(self):
        response = client.get("/api/export/auction/9999")
        assert response.status_code == 404

    def test_export_inventory(self):
        response = client.get("/api/export/inventory")
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

    def test_export_dashboard_summary(self):
        response = client.get("/api/export/dashboard/summary?year=2026")
        assert response.status_code == 200

    def test_export_tax_summary(self):
        response = client.get("/api/export/tax-summary?year=2026")
        assert response.status_code == 200

    def test_export_columns_present(self):
        resp = client.post("/api/auctions/", json={"name": "Col Test", "date": "2026-03-25"})
        aid = resp.json()["id"]
        client.post(f"/api/auctions/{aid}/items", json={"name": "Item", "hammer_price": 10.0})
        response = client.get(f"/api/export/auction/{aid}")
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Net Profit" in headers
        assert "ROI %" in headers


# ── API endpoint tests for new routers ───────────────────────────────────────

class TestAuctionHouseAPI:
    def test_list_auction_houses(self):
        resp = client.get("/api/auction-houses/")
        assert resp.status_code == 200
        data = resp.json()
        assert len(data) >= 6  # 6 seeded presets

    def test_get_provinces(self):
        resp = client.get("/api/auction-houses/provinces")
        assert resp.status_code == 200
        data = resp.json()
        assert "ON" in data
        assert data["ON"]["rate"] == 13.0

    def test_preview_buy_cost(self):
        resp = client.post("/api/auction-houses/preview", json={
            "hammer": 50.0, "buyer_premium_pct": 16.0,
            "handling_fee_mode": "flat", "handling_fee_flat": 1.50,
            "tax_pct": 13.0, "tax_applies_to": "subtotal",
            "payment_method": "etransfer",
        })
        assert resp.status_code == 200
        assert resp.json()["total_buy_cost"] > 50.0


class TestShippingPresetsAPI:
    def test_list_shipping_presets(self):
        resp = client.get("/api/shipping-presets/")
        assert resp.status_code == 200
        data = resp.json()
        assert len(data) >= 9  # 9 seeded presets


class TestTemplatesAPI:
    def test_list_templates(self):
        resp = client.get("/api/templates/")
        assert resp.status_code == 200

    def test_create_and_delete_profile(self):
        resp = client.post("/api/templates/", json={
            "item_name": "Test Dyson", "category": "Electronics",
            "sale_channel": "ebay", "profile_type": "auction",
        })
        assert resp.status_code == 200
        tmpl_id = resp.json()["id"]
        del_resp = client.delete(f"/api/templates/{tmpl_id}")
        assert del_resp.status_code == 200


class TestCurrencyAPI:
    def test_get_currency_rate(self):
        resp = client.get("/api/currency/rate")
        assert resp.status_code == 200
        data = resp.json()
        assert "cad_to_usd" in data
        assert data["cad_to_usd"] > 0


# ── Bug Fix 2 — Custom Tax Override ──────────────────────────────────────────

from app.buy_cost import build_config_with_settings, get_tax_label


class TestCustomTaxOverride:
    def test_custom_tax_overrides_province(self):
        settings = {"use_custom_tax": "1", "custom_tax_pct": "8.5"}
        config = build_config_with_settings(settings)
        result = calculate_buy_cost(100.00, config, "cash")
        # Encore defaults: premium=16, handling=1.50, subtotal=117.50, tax=117.50*0.085=9.9875
        assert result["tax"] == pytest.approx(9.99, abs=0.02)

    def test_province_tax_used_when_custom_disabled(self):
        settings = {"use_custom_tax": "0", "province": "ON"}
        config = build_config_with_settings(settings)
        result = calculate_buy_cost(100.00, config, "cash")
        assert result["tax"] == pytest.approx(15.28, abs=0.02)  # 13% of 117.50 = 15.275 → 15.28

    def test_tax_label_custom(self):
        settings = {"use_custom_tax": "1", "custom_tax_label": "State/Local Tax"}
        assert get_tax_label(settings) == "State/Local Tax"

    def test_tax_label_province(self):
        settings = {"use_custom_tax": "0", "province": "ON"}
        assert get_tax_label(settings) == "HST"

    def test_tax_label_empty_custom(self):
        settings = {"use_custom_tax": "1", "custom_tax_label": ""}
        assert get_tax_label(settings) == "Custom Tax"


# ── Bug Fix 3 — Export Fixes ─────────────────────────────────────────────────

class TestExportFixes:
    def test_export_empty_auction_returns_file(self):
        resp = client.post("/api/auctions/", json={"name": "Empty Auction", "date": "2026-03-27"})
        aid = resp.json()["id"]
        response = client.get(f"/api/export/auction/{aid}")
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
        # Verify the file has content (headers + "No items" row)
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "No items in this auction"

    def test_export_nonexistent_auction_returns_404(self):
        response = client.get("/api/export/auction/999999")
        assert response.status_code == 404


# ── Bug Fix 4 — Profit Recalculation ─────────────────────────────────────────

class TestProfitRecalc:
    def _create_sold_item(self):
        resp = client.post("/api/auctions/", json={"name": "Recalc Test", "date": "2026-03-27"})
        aid = resp.json()["id"]
        item_resp = client.post(f"/api/auctions/{aid}/items", json={
            "name": "Sold Widget", "hammer_price": 30.0, "status": "sold",
            "sold_price": 80.0, "sell_date": "2026-03-27", "sale_channel": "ebay",
        })
        return item_resp.json()

    def test_profit_calculated_on_create(self):
        item = self._create_sold_item()
        assert item["net_profit"] is not None
        assert item["roi_pct"] is not None

    def test_profit_resets_when_status_set_to_unlisted(self):
        item = self._create_sold_item()
        r = client.put(f"/api/auctions/items/{item['id']}", json={"status": "unlisted", "sold_price": None})
        assert r.status_code == 200
        data = r.json()
        assert data["net_profit"] is None
        assert data["roi_pct"] is None

    def test_profit_recalculates_on_price_change(self):
        item = self._create_sold_item()
        original_profit = item["net_profit"]
        r = client.put(f"/api/auctions/items/{item['id']}", json={"sold_price": 200.00})
        assert r.status_code == 200
        new_profit = r.json()["net_profit"]
        assert new_profit is not None
        assert new_profit != original_profit
        assert new_profit > original_profit  # Higher sell = higher profit

    def test_profit_recalculates_on_hammer_price_change(self):
        item = self._create_sold_item()
        original_cost = item["buy_cost_total"]
        r = client.put(f"/api/auctions/items/{item['id']}", json={"hammer_price": 1.00})
        assert r.status_code == 200
        data = r.json()
        assert data["buy_cost_total"] != original_cost
        assert data["buy_cost_total"] < original_cost  # Lower hammer = lower cost


# ── Bug Fix 5 — Alert Endpoint ───────────────────────────────────────────────

class TestAlertEndpoint:
    def test_alert_returns_200(self):
        r = client.post("/api/calculator/alert", json={
            "hammer_price": 10.00, "sell_price": 25.00, "channel": "ebay",
            "shipping_cost": 5.00, "payment_method": "etransfer", "target_margin_pct": 30.0,
        })
        assert r.status_code == 200
        assert "level" in r.json()

    def test_alert_red_below_cost(self):
        r = client.post("/api/calculator/alert", json={
            "hammer_price": 50.00, "sell_price": 5.00, "channel": "ebay",
            "shipping_cost": 0, "payment_method": "etransfer",
        })
        assert r.json()["level"] == "red"

    def test_alert_green_high_margin(self):
        r = client.post("/api/calculator/alert", json={
            "hammer_price": 10.00, "sell_price": 200.00, "channel": "ebay",
            "shipping_cost": 0, "payment_method": "etransfer", "target_margin_pct": 30.0,
        })
        assert r.json()["level"] == "green"

    def test_alert_with_kijiji_channel(self):
        r = client.post("/api/calculator/alert", json={
            "hammer_price": 10.00, "sell_price": 25.00, "channel": "kijiji",
            "shipping_cost": 0, "payment_method": "etransfer",
        })
        assert r.status_code == 200
        assert "level" in r.json()


# ── Product Profiles CRUD ────────────────────────────────────────────────────

class TestProductProfiles:
    def test_auction_profile_has_no_buy_price(self):
        r = client.post("/api/templates/", json={
            "item_name": "Dyson V8", "profile_type": "auction",
            "category": "Vacuums", "sale_channel": "ebay",
            "promoted_listing_pct": 5.0, "target_margin_pct": 35.0,
        })
        assert r.status_code == 200
        assert r.json()["fixed_buy_price"] is None
        assert r.json()["profile_type"] == "auction"

    def test_fixed_profile_stores_buy_price(self):
        r = client.post("/api/templates/", json={
            "item_name": "AirPods Pro", "profile_type": "fixed",
            "category": "Electronics", "sale_channel": "ebay",
            "fixed_buy_price": 45.00, "typical_sell_price": 120.00,
        })
        assert r.status_code == 200
        data = r.json()
        assert data["fixed_buy_price"] == 45.00
        assert data["typical_sell_price"] == 120.00
        assert data["profile_type"] == "fixed"

    def test_item_name_required(self):
        r = client.post("/api/templates/", json={
            "item_name": "", "profile_type": "auction",
        })
        assert r.status_code == 422

    def test_item_name_required_missing(self):
        r = client.post("/api/templates/", json={
            "profile_type": "auction", "category": "Books",
        })
        assert r.status_code == 422

    def test_get_profile_by_id(self):
        r = client.post("/api/templates/", json={
            "item_name": "Test Get", "category": "Electronics",
            "sale_channel": "ebay", "promoted_listing_pct": 5.0,
        })
        tid = r.json()["id"]
        r2 = client.get(f"/api/templates/{tid}")
        assert r2.status_code == 200
        assert r2.json()["item_name"] == "Test Get"

    def test_get_profile_not_found(self):
        r = client.get("/api/templates/999999")
        assert r.status_code == 404

    def test_update_profile(self):
        r = client.post("/api/templates/", json={
            "item_name": "Update Me", "category": "Books", "sale_channel": "ebay",
        })
        tid = r.json()["id"]
        r2 = client.put(f"/api/templates/{tid}", json={"target_margin_pct": 40.0})
        assert r2.status_code == 200
        assert r2.json()["target_margin_pct"] == 40.0

    def test_delete_profile(self):
        r = client.post("/api/templates/", json={"item_name": "Delete Me", "category": "Toys"})
        tid = r.json()["id"]
        del_r = client.delete(f"/api/templates/{tid}")
        assert del_r.status_code == 200
        r2 = client.get(f"/api/templates/{tid}")
        assert r2.status_code == 404

    def test_profile_list_returns_both_types(self):
        client.post("/api/templates/", json={
            "item_name": "Auction Item", "profile_type": "auction",
        })
        client.post("/api/templates/", json={
            "item_name": "Fixed Item", "profile_type": "fixed",
            "fixed_buy_price": 10.0,
        })
        r = client.get("/api/templates/")
        types = [p["profile_type"] for p in r.json()]
        assert "auction" in types
        assert "fixed" in types

    def test_ebay_settings_saved(self):
        r = client.post("/api/templates/", json={
            "item_name": "eBay Profile", "profile_type": "auction",
            "ebay_category": "Collectibles", "ebay_store_toggle": 1,
            "top_rated_toggle": 1, "insertion_fee_toggle": 1,
        })
        assert r.status_code == 200
        data = r.json()
        assert data["ebay_category"] == "Collectibles"
        assert data["ebay_store_toggle"] == 1
        assert data["top_rated_toggle"] == 1
        assert data["insertion_fee_toggle"] == 1
